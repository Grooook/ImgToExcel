import os

import easygui
from openpyxl.styles import PatternFill, colors
from openpyxl import Workbook
from PIL import Image
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk
import threading


def open_explorer():
    msg = 'only .jpg/.png'
    title = 'Choose File'
    filetypes = ['*.jpg']
    while True:
        file = easygui.fileopenbox(msg, title, filetypes=filetypes)
        if file.endswith('.jpg') or file.endswith('.png'):
            break
    return file


def open_resize_box(size):
    msg = 'Give new sizes'
    title = "Resize this image"
    field_names = [f'Width (default = {size[0]})', f'Heigth (default = {size[1]})']
    field_values = easygui.multenterbox(msg, title, field_names)

    while True:
        if field_values is None:
            break
        error_msg = ""
        for i in range(len(field_names)):
            if field_values[i].strip() == "":
                error_msg = error_msg + ('"%s" is a required field.\n\n' % field_names[i])

            elif not field_values[i].isnumeric():
                error_msg = error_msg + ('"%s" must be a positive number.\n\n' % field_names[i])

            elif field_values[i].isnumeric():
                if int(field_values[i]) < 50:
                    error_msg = error_msg + ('"%s" must be more than 50.\n\n' % field_names[i])

                if i == 0 and int(field_values[i]) > size[0]:
                    error_msg = error_msg + ('"%s" must be less than default width.\n\n' % field_names[i])

                if i == 1 and int(field_values[i]) > size[1]:
                    error_msg = error_msg + ('"%s" must be less than default heigth.\n\n' % field_names[i])

        if error_msg == "":
            break
        field_values = easygui.multenterbox(error_msg, title, field_names, field_values)
    return [int(i) for i in field_values]


def progress_bar():
    window.mainloop()


def bar_update():
    bar['value'] += 1


window = tk.Tk()
window.title("Tkinter Progressbar")
window.geometry('200x50')
window.withdraw()
style = ttk.Style()
style.theme_use('default')
style.configure("grey.Horizontal.TProgressbar", background='blue')
bar = ttk.Progressbar(window, length=200, style='grey.Horizontal.TProgressbar')
text = tk.StringVar()
progress_text = tk.Label(window, textvariable=text)
bar['value'] = 0
progress_text.grid(column=0, row=1)
bar.grid(column=0, row=0)


def read_img():
    img = Image.open(open_explorer())
    new_size = open_resize_box(img.size)
    img = img.resize((new_size[0], new_size[1]))
    img.save('resized.jpg')
    del img
    img = Image.open('resized.jpg')
    pix = img.load()
    x, y = img.size
    hex_list = [[i for i in range(y)] for _ in range(x)]
    for i in range(x):
        for j in range(y):
            rgb = (pix[i, j])
            hex_color = ('%02x%02x%02x' % rgb)
            hex_list[i][j] = hex_color
    os.remove('resized.jpg')
    fill(hex_list, x, y)


def fill(hex_list, x, y):
    workbook = Workbook()
    sheet = workbook.active

    for i in range(x):
        sheet.column_dimensions[get_column_letter(i + 1)].width = 3
    thread = threading.Thread(target=fill_excel, args=(hex_list, x, y, sheet))
    thread.start()
    window.deiconify()
    window.mainloop()
    file_name = easygui.filesavebox(title="Exporting ascii", filetypes=".jpg")
    if not file_name.endswith('.xlsx') or file_name.endswith('.csv'):
        file_name += '.xlsx'

    workbook.save(filename=file_name)


def fill_excel(hex_list, x, y, sheet):
    counter = 0
    for i in range(x):
        for j in range(y):
            counter += 1
            my_color = colors.Color(rgb=hex_list[i][j])
            sheet.cell(row=j + 1, column=i + 1).fill = PatternFill(fgColor=my_color, fill_type='solid')
            if counter / (x * y) * 100 > bar['value']:
                bar_update()
                text.set(f'{counter  * 100 // (x * y)} / 100%')
    window.destroy()


if __name__ == '__main__':
    read_img()
